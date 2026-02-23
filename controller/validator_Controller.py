# ============================================================================
# controllers/validador_controller.py
# CONTROLADOR - Maneja la lógica entre Vista y Modelo (MODIFICADO)
# ============================================================================

import os
import threading
import time
import subprocess
import platform
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook

from model.link_validator import LinkValidator
from view.validador_view import ValidadorView
from logger import get_logger
from errors import MensajesError, registrar_error, mostrar_error
from config.constants import (
    DELAY_DEFAULT,
    VALOR_EXCEL_VALIDO,
    VALOR_EXCEL_NO_VALIDO,
    VALOR_EXCEL_VALIDAR,
    EMOJI_VALIDO,
    EMOJI_NO_VALIDO,
    EMOJI_VALIDAR,
    EMOJI_ARCHIVO,
    EMOJI_HOJA,
    EMOJI_INICIO,
    EMOJI_ESTADO,
    EMOJI_CUIDADO,
    EMOJI_CONTINUAR,
    EMOJI_CONFIGURACION,
    EMOJI_GUARDADO,
    EMOJI_INICIO
)


class ValidadorController:    
    def __init__(self, root):
        self.root = root
        self.logger = get_logger(guardar_en_archivo=True, ruta_archivo="logs_validacion.txt")
        
        # Crear Modelo y Vista
        self.modelo = LinkValidator()
        self.vista = ValidadorView(root)
        
        # Variables de estado del controlador
        self.excel_path = None
        self.df = None
        self.columnas = []
        self.validacion_corriendo = False
        self.detener_validacion = False
        self.pausar_validacion = False  
        
        # Conectar eventos de la Vista con métodos del Controlador
        self.conectar_eventos()
        
        self.configurar_drag_drop()
        
        self.logger.success("Aplicación iniciada correctamente")
    
    # ========================================================================
    # CONFIGURACIÓN DE EVENTOS
    # ========================================================================
    
    def conectar_eventos(self):
        self.vista.vincular_click_archivo(self.seleccionar_archivo)
        self.vista.vincular_boton_ejecutar(self.iniciar_validacion)
        self.vista.vincular_boton_ver_logs(self.abrir_logs)
        self.vista.vincular_boton_pausar(self.pausar_reanudar_validacion)
        self.vista.vincular_boton_detener(self.detener_validacion_manual)
    
    def configurar_drag_drop(self):
        try:
            from tkinterdnd2 import TkinterDnD
            if isinstance(self.root, TkinterDnD.Tk):
                self.vista.vincular_drag_drop(self.archivo_arrastrado)
                self.logger.success("Drag & Drop habilitado")
            else:
                self.logger.warning("TkinterDnD no inicializado correctamente")
        except ImportError:
            self.logger.warning("Drag & Drop no disponible (instala tkinterdnd2)")
        except Exception as e:
            self.logger.warning(f"Error Drag & Drop: {e}")
    
    # ========================================================================
    # ABRIR ARCHIVO DE LOGS
    # ========================================================================
    
    def abrir_logs(self):
       
        ruta_logs = "logs_validacion.txt"
        if not os.path.exists(ruta_logs):
            messagebox.showinfo(
                "Logs no disponibles",
                "Aún no se han generado logs.\n\nLos logs se crean al cargar un archivo o iniciar una validación."
            )
            return
        
        try:
            # Abrir con el programa predeterminado según el sistema operativo
            if platform.system() == 'Windows':
                os.startfile(ruta_logs)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.call(['open', ruta_logs])
            else:  # Linux
                subprocess.call(['xdg-open', ruta_logs])
            
            self.logger.info(f"{EMOJI_HOJA} Archivo de logs abierto")
            
        except Exception as e:
            msg = MensajesError.error_inesperado(type(e).__name__, str(e), "Abrir archivo de logs")
            registrar_error(self.logger, msg)
            mostrar_error(msg, messagebox)
    
    # ========================================================================
    # MANEJO DE ARCHIVOS EXCEL
    # ========================================================================
    
    def seleccionar_archivo(self, event=None):
        filepath = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filepath:
            self.cargar_archivo(filepath)
    
    def archivo_arrastrado(self, event):
        filepath = event.data
        if filepath.startswith('{'):
            filepath = filepath[1:-1]
        self.cargar_archivo(filepath)
    
    def cargar_archivo(self, filepath):
        try:
            nombre_archivo = os.path.basename(filepath)
            self.logger.info(f"{EMOJI_ARCHIVO} Intentando cargar archivo: {nombre_archivo}")
            
            # Leer archivo Excel y obtener hojas
            xls = pd.ExcelFile(filepath)
            hojas = xls.sheet_names
            
            # Leer la primera hoja
            self.df = pd.read_excel(filepath, sheet_name=0)
            self.excel_path = filepath
            self.columnas = self.df.columns.tolist()
            
            # Actualizar la Vista
            self.vista.mostrar_archivo_cargado(
                nombre_archivo,
                len(self.df),
                len(self.columnas),
                hojas
            )
            
            # Log de éxito usando el logger
            self.logger.log_carga_archivo(nombre_archivo, len(self.df), len(self.columnas), hojas)
            
        except FileNotFoundError as e:
            # Usar mensaje centralizado
            msg = MensajesError.archivo_no_encontrado(filepath, str(e))
            registrar_error(self.logger, msg)
            self.vista.mostrar_error_carga()
            mostrar_error(msg, messagebox)
            
        except PermissionError as e:
            # Usar mensaje centralizado
            msg = MensajesError.archivo_bloqueado(nombre_archivo, str(e))
            registrar_error(self.logger, msg)
            self.vista.mostrar_error_carga()
            mostrar_error(msg, messagebox)
            
        except ValueError as e:
            # Usar mensaje centralizado
            msg = MensajesError.archivo_corrupto(nombre_archivo, str(e))
            registrar_error(self.logger, msg)
            self.vista.mostrar_error_carga()
            mostrar_error(msg, messagebox)
            
        except Exception as e:
            # Usar mensaje centralizado para error inesperado
            msg = MensajesError.error_inesperado(
                type(e).__name__,
                str(e),
                f"Cargando archivo: {nombre_archivo}"
            )
            registrar_error(self.logger, msg)
            self.vista.mostrar_error_carga()
            mostrar_error(msg, messagebox)
    
    # ========================================================================
    # VALIDACIÓN DE CONFIGURACIÓN
    # ========================================================================
    
    def validar_configuracion(self):
        if not self.excel_path or self.df is None:
            messagebox.showwarning("Advertencia", "Primero debes cargar un archivo Excel")
            return False
        
        # Obtener valores de la Vista
        columna = self.vista.obtener_columna()
        
        # Validar columna
        if not columna:
            messagebox.showwarning("Advertencia", "Debes ingresar la letra de la columna")
            return False
        
        if not columna.isalpha():
            messagebox.showerror("Error", "La columna debe contener solo letras (ej: A, B, C, AA)")
            return False
        
        # Validar filas
        try:
            fila_ini = self.vista.obtener_fila_inicio()
            fila_fin = self.vista.obtener_fila_fin()
            
            if fila_ini < 1:
                messagebox.showerror("Error", "La fila inicial debe ser mayor a 0")
                return False
                
            if fila_ini > fila_fin:
                messagebox.showerror("Error", "La fila inicial debe ser menor o igual a la final")
                return False
                
            # Verificar que la columna exista en el archivo
            col_idx = self.modelo.letra_a_indice(columna)
            if col_idx >= len(self.df.columns):
                messagebox.showerror(
                    "Error", 
                    f"La columna '{columna}' está fuera del rango.\n"
                    f"El archivo tiene {len(self.df.columns)} columnas."
                )
                return False
                
        except ValueError:
            messagebox.showerror("Error", "Las filas deben ser números enteros válidos")
            return False
        
        return True
    
    # ========================================================================
    # INICIO Y EJECUCIÓN DE VALIDACIÓN
    # ========================================================================
    
    def iniciar_validacion(self):
        # Verificar si ya hay una validación corriendo
        if self.validacion_corriendo:
            messagebox.showwarning("Advertencia", "Ya hay una validación en curso")
            return
        
        # Validar configuración
        if not self.validar_configuracion():
            return
        
        # Marcar que la validación está corriendo
        self.validacion_corriendo = True
        self.detener_validacion = False
        self.pausar_validacion = False
        self.vista.deshabilitar_boton_ejecutar()
        
        # Ejecutar en hilo separado para no bloquear la interfaz
        thread = threading.Thread(target=self.ejecutar_validacion, daemon=True)
        thread.start()
    
    def pausar_reanudar_validacion(self):
        if not self.validacion_corriendo:
            return
        
        if self.pausar_validacion:
            # Reanudar
            self.pausar_validacion = False
            self.vista.cambiar_boton_reanudar_a_pausar()
            self.logger.info(f"{EMOJI_CONTINUAR} Validación reanudada")
        else:
            # Pausar
            self.pausar_validacion = True
            self.vista.cambiar_boton_pausar_a_reanudar()
            self.logger.warning("⏸ Validación pausada")
            self.vista.actualizar_progreso("⏸ PAUSADO - Click REANUDAR para continuar", self.vista.progreso_bar['value'])
    
    def detener_validacion_manual(self):
        if not self.validacion_corriendo:
            return
        
        respuesta = messagebox.askyesno(
            "Confirmar detención",
            "¿Estás seguro de que quieres detener la validación?\n\n"
            "Los resultados procesados hasta ahora se perderán."
        )
        
        if respuesta:
            self.detener_validacion = True
            self.logger.warning(f"{EMOJI_CUIDADO} Usuario solicitó detener la validación")
            self.vista.actualizar_progreso("Deteniendo validación...", 0)
    
    def ejecutar_validacion(self):
        try:
            # Obtener valores de la Vista
            columna = self.vista.obtener_columna()
            fila_ini = self.vista.obtener_fila_inicio()
            fila_fin = self.vista.obtener_fila_fin()
            hoja = self.vista.obtener_hoja()
            
            delay = DELAY_DEFAULT
            
            self.logger.info(f"{EMOJI_CONFIGURACION} Configuración de validación:")
            self.logger.info(f"{EMOJI_ARCHIVO}  → Archivo: {os.path.basename(self.excel_path)}")
            self.logger.info(f"   → Hoja: {hoja}")
            self.logger.info(f"   → Columna: {columna}")
            self.logger.info(f"   → Rango: Filas {fila_ini} a {fila_fin}")
            
            # Cargar la hoja seleccionada (sin encabezados)
            self.df = pd.read_excel(self.excel_path, sheet_name=hoja, header=None)
            
            # Calcular índices de pandas (empiezan en 0)
            idx_inicio_pandas = fila_ini - 1
            idx_fin_pandas = fila_fin
            
            # Convertir letra de columna a índice
            col_idx = self.modelo.letra_a_indice(columna)
            
            # Filtrar el DataFrame
            df_filtrado = self.df.iloc[idx_inicio_pandas:idx_fin_pandas].copy()
            total = len(df_filtrado)
            
            if total == 0:
                # Usar mensaje centralizado
                msg = MensajesError.rango_sin_datos(fila_ini, fila_fin)
                self.logger.warning(f"{EMOJI_VALIDAR} {msg['que_paso']}")
                self.logger.warning(f"   → {msg['por_que']}")
                messagebox.showwarning(msg['titulo'], MensajesError.formatear_para_popup(msg))
                return
            
            # Log de inicio usando el logger
            self.logger.log_inicio_validacion(hoja, columna, fila_ini, fila_fin, total, delay)
            
            # Configurar barra de progreso
            self.vista.configurar_progreso_maximo(total)
            
            # Registrar tiempo de inicio
            inicio = time.time()
            
            # Preparar lista de URLs con sus filas de Excel
            urls_a_validar_con_fila = []
            for idx_pandas, row in df_filtrado.iterrows():
                fila_excel = idx_pandas + 1
                
                if col_idx < len(row):
                    urls_a_validar_con_fila.append((fila_excel, row[col_idx]))
                else:
                    # Usar mensaje centralizado
                    msg = MensajesError.columna_fuera_limites(columna, fila_excel, len(row))
                    self.logger.warning(f"{EMOJI_VALIDAR} {msg['que_paso']}")
                    self.logger.warning(f"   → {msg['por_que']}")
                    self.logger.warning(f"   → Solución: {msg['como_resolver']}")
            
            if not urls_a_validar_con_fila:
                self.logger.error(f"{EMOJI_NO_VALIDO} ERROR: No se encontraron URLs válidas en el rango especificado")
                raise ValueError("No se encontraron URLs válidas en el rango especificado")
            
            # VALIDAR EN LOTE usando el Modelo
            self.logger.info(f"{EMOJI_INICIO} Iniciando validación de {len(urls_a_validar_con_fila)} URLs...")
            resultados = []
            
            for idx, (fila_excel, url) in enumerate(urls_a_validar_con_fila, 1):
                # Verificar si se solicitó detener
                if self.detener_validacion:
                    self.logger.warning(f"{EMOJI_CUIDADO} Validación detenida por el usuario")
                    self.logger.warning(f"   → Procesadas: {idx-1}/{len(urls_a_validar_con_fila)} URLs")
                    messagebox.showinfo(
                        "Validación detenida",
                        f"Validación detenida por el usuario.\n\n"
                        f"URLs procesadas: {idx-1}/{len(urls_a_validar_con_fila)}\n\n"
                        f"Los resultados NO se guardaron."
                    )
                    return
                
                # Verificar si está pausado (esperar hasta que reanude)
                while self.pausar_validacion:
                    time.sleep(0.5)  # Esperar medio segundo
                    if self.detener_validacion:  # Por si detiene mientras está pausado
                        return
                try:
                    disp_url = str(url) if not pd.isna(url) else "<VACÍA>"
                except:
                    disp_url = "<VACÍA>"
                short_disp = disp_url[:60] + "..." if len(disp_url) > 60 else disp_url
                
                self.vista.actualizar_progreso(f"Analizando {idx}/{len(urls_a_validar_con_fila)} (Fila {fila_excel}): {short_disp}", idx - 1)

                # Validar URL
                resultado = self.modelo.validar_url(url, delay)
                resultado['fila_excel'] = fila_excel
                resultados.append(resultado)
                
                # Callback de progreso
                self.callback_progreso(url, idx, len(urls_a_validar_con_fila), resultado, fila_excel)
            
            # Calcular tiempo total
            tiempo_total = time.time() - inicio
            
            # Guardar resultados en Excel
            self.logger.info(f"{EMOJI_GUARDADO} Guardando resultados en el archivo Excel...")
            contadores = self.guardar_resultados(hoja, resultados)
            
            # Log final usando el logger
            self.logger.log_fin_validacion(
                tiempo_total,
                contadores['validos'],
                contadores['no_validos'],
                contadores['validar']
            )
            
            # Mostrar mensaje de éxito
            messagebox.showinfo(
                f"{EMOJI_VALIDO} Validación Completada", 
                f"Validación finalizada exitosamente\n\n"
                f"{EMOJI_ESTADO} Procesadas: {total} URLs\n"
                f"{EMOJI_VALIDO}  Válidas: {contadores['validos']}\n"
                f"{EMOJI_NO_VALIDO}  No válidas: {contadores['no_validos']}\n"
                f"{EMOJI_VALIDAR}  Validar manualmente: {contadores['validar']}\n\n"
                f"{EMOJI_GUARDADO} Resultados guardados en:\n{os.path.basename(self.excel_path)}"
            )
            
        except FileNotFoundError as e:
            msg = MensajesError.archivo_no_encontrado(self.excel_path, str(e))
            registrar_error(self.logger, msg)
            mostrar_error(msg, messagebox)
            
        except PermissionError as e:
            msg = MensajesError.error_guardar_resultados(os.path.basename(self.excel_path), str(e))
            mostrar_error(msg, messagebox)
            
        except ValueError as e:
            msg = MensajesError.configuracion_invalida(str(e))
            registrar_error(self.logger, msg)
            mostrar_error(msg, messagebox)
            
        except Exception as e:
            msg = MensajesError.error_inesperado(type(e).__name__, str(e), "Ejecución de validación")
            registrar_error(self.logger, msg)
            mostrar_error(msg, messagebox)
        
        finally:
            # Siempre resetear estado al finalizar
            self.logger.info(f"{EMOJI_INICIO} Finalizando proceso de validación...")
            self.validacion_corriendo = False
            self.detener_validacion = False
            self.pausar_validacion = False
            self.vista.habilitar_boton_ejecutar()
            self.vista.deshabilitar_boton_pausar()
            self.vista.deshabilitar_boton_detener()
            self.vista.resetear_progreso()
    
    # ========================================================================
    # CALLBACK DE PROGRESO
    # ========================================================================
    
    def callback_progreso(self, url, idx, total, resultado, fila_excel):
        try:
            display_url = str(url) if not pd.isna(url) else "<VACÍA>"
        except Exception:
            display_url = "<VACÍA>"

        # Acortar URL si es muy larga
        short_display = display_url[:60] + "..." if len(display_url) > 60 else display_url

        # Actualizar Vista
        fila_info = f"(Fila {fila_excel})"
        texto_progreso = f"Validando {idx}/{total} {fila_info}: {short_display}"
        self.vista.actualizar_progreso(texto_progreso, idx)

        # Solo hacer log si hay un estado definido (no celdas vacías o texto inválido)
        estado = resultado.get('estado')
        if estado is not None:
            # Log usando el logger
            self.logger.log_validacion(estado, fila_excel, display_url, idx, total)
        else:
            # Para celdas vacías o texto no-URL, log más simple
            detalles = resultado.get('detalles', '')
            if detalles == 'Celda vacía':
                self.logger.debug(f"[{idx}/{total}] Fila {fila_excel}: Celda vacía (omitida)")
            elif 'no es una URL' in detalles:
                # Acortar el texto si es muy largo
                texto_corto = display_url[:30] + "..." if len(display_url) > 30 else display_url
                self.logger.debug(f"[{idx}/{total}] Fila {fila_excel}: Texto no válido '{texto_corto}' (omitido)")
    
    # ========================================================================
    # GUARDADO DE RESULTADOS EN EXCEL
    # ========================================================================
    
    def guardar_resultados(self, hoja, resultados):
        try:
            self.logger.info(f"{EMOJI_GUARDADO} Abriendo archivo para guardar resultados...")
            wb = load_workbook(self.excel_path)
            ws = wb[hoja]
            
            # Obtener columna de resultados
            col_resultado = self.vista.obtener_columna_resultado()
            col_resultado_idx = self.modelo.letra_a_indice(col_resultado) + 1
            
            self.logger.info(f"{EMOJI_HOJA} Escribiendo resultados en columna '{col_resultado}'...")
            
            # Contadores para el resumen
            contadores = {'validos': 0, 'no_validos': 0, 'validar': 0}
            
            # Escribir resultados
            for res in resultados:
                fila_excel = res.get('fila_excel')
                estado = res.get('estado')
                
                if fila_excel is None:
                    continue
                
                # Si no hay estado (celda vacía), no escribir nada
                if estado is None:
                    ws.cell(row=fila_excel, column=col_resultado_idx, value=None)
                    continue
                
                # Determinar valor a escribir usando constantes
                if estado == 'valido':
                    valor = VALOR_EXCEL_VALIDO
                    contadores['validos'] += 1
                elif estado == 'validar':
                    valor = VALOR_EXCEL_VALIDAR
                    contadores['validar'] += 1
                else:  # no_valido
                    valor = VALOR_EXCEL_NO_VALIDO
                    contadores['no_validos'] += 1
                
                # Escribir en la celda
                ws.cell(row=fila_excel, column=col_resultado_idx, value=valor)
            
            # Guardar y cerrar
            self.logger.info(f"{EMOJI_GUARDADO} Guardando cambios en el archivo...")
            wb.save(self.excel_path)
            wb.close()
            
            self.logger.success(f"{EMOJI_VALIDO} Resultados guardados exitosamente en columna '{col_resultado}'")
            
            return contadores
            
        except PermissionError as e:
            msg = MensajesError.error_guardar_resultados(os.path.basename(self.excel_path), str(e))
            registrar_error(self.logger, msg)
            raise PermissionError(f"El archivo está bloqueado: {str(e)}")
            
        except Exception as e:
            msg = MensajesError.error_inesperado(type(e).__name__, str(e), "Guardando resultados")
            registrar_error(self.logger, msg)
            raise